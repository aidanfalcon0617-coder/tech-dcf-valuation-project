"""
build_dcf_model.py  (tech DCF project)

Reads the cleaned annual financials from clean_data.py and writes a real,
editable DCF model as an .xlsx workbook -- one sheet per company plus a
Summary sheet, with the whole projection built out of live Excel formulas
(not hardcoded numbers), so changing WACC, terminal growth, the starting
growth rate, or the FCF margin recalculates everything downstream.

Modeling approach and why
--------------------------
The textbook unlevered-FCF buildup is EBIT*(1-tax) + D&A - capex - change in
NWC. That buildup breaks down for most of this company set: heavy
stock-based comp (a non-cash expense) pushes GAAP operating income negative
for half of them in recent years (CRWD, DDOG, SNOW, TEAM, ZS all show
negative operating_income in dcf_financials.csv), and near-zero pretax
income makes the reported effective tax rate swing wildly and
meaninglessly (Salesforce 68%, Workday -288%, CrowdStrike -27%, etc --
see effective_tax_rate in the cleaned data). Feeding either of those into a
DCF produces nonsense.

What's actually stable for this group is the relationship between revenue
and free cash flow: FCF margins (free_cash_flow / revenue) cluster in a
believable 20-40% band and don't swing sign the way EBIT does, because free
cash flow already nets out the SBC add-back, capex, and working capital
changes as actually realized in cash. So the model here is:

    revenue_t = revenue_(t-1) * (1 + growth_t)
    growth_t  = fades linearly from a starting rate to the terminal rate
                over the projection horizon (matches growth_N = terminal
                growth exactly, so the Gordon-growth terminal value is
                internally consistent)
    FCF_t     = revenue_t * fcf_margin        (fcf_margin held flat)

This is a standard simplification for high-growth SaaS DCFs where the
GAAP income statement is dominated by non-cash comp. It trades some
textbook rigor for a model that isn't garbage-in-garbage-out on this
specific company set -- and it's easy to explain and defend, which matters
more than superficial textbook completeness (see README).

Assumption defaults (all four are editable input cells in the workbook):
  - WACC: 10% for every company. The README calls for "a single assumed
    WACC" -- a deliberate simplification, not company-specific CAPM/beta.
  - Terminal growth rate: 3% (long-run nominal GDP-ish), same for everyone.
  - Starting revenue growth rate: each company's own most recent single-year
    YoY revenue growth (not a multi-year CAGR -- growth is decelerating for
    every company in this set, so a backward-looking average would overstate
    the actual current run rate the fade should start from).
  - FCF margin: average of the two most recent fiscal years' FCF/revenue,
    to smooth one noisy year without averaging over a period where the
    margin was structurally different.

Net debt (for the enterprise-to-equity bridge) uses the most recent
fiscal year's total_debt and cash_and_equivalents. HubSpot's total_debt is
NULL for its most recent year in the cleaned data -- not because the field
was never reported (it reported real convertible-note balances in
2017-2021 and 2024), but because that specific year's 10-K genuinely
carries no debt tag, most likely between note issuances. Treated as $0 for
that year, exactly like clean_data.py already treats a missing debt
component as $0 when only one side of short/long-term is reported.

Usage:
    python build_dcf_model.py
"""

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DATA_DIR = PROJECT_ROOT / "data" / "processed"
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_PATH = OUTPUT_DIR / "tech_dcf_valuation.xlsx"

DEFAULT_WACC = 0.10
DEFAULT_TERMINAL_GROWTH = 0.03
PROJECTION_YEARS = 5
FCF_MARGIN_LOOKBACK_YEARS = 2

# Sensitivity grid axes: WACC down the rows, terminal growth across the columns,
# centered on the DEFAULT_WACC / DEFAULT_TERMINAL_GROWTH base case.
SENSITIVITY_WACC_VALUES = [0.08, 0.09, 0.10, 0.11, 0.12]
SENSITIVITY_TG_VALUES = [0.02, 0.025, 0.03, 0.035, 0.04]

# One-time market snapshot (price, shares outstanding in millions) as of
# market close, used to seed the Summary sheet's "current price" / "current
# shares" input cells -- both stay editable so the comparison doesn't go
# stale. Cross-checked against a second source (Google Finance) for CRM and
# CRWD; CRWD's share count reflects a stock split in the week before this
# snapshot, which is why it doesn't match this project's FY2026 diluted
# share count (that 10-K predates the split) -- see the README for why the
# comparison below uses market cap, not price per share, for exactly that
# reason.
MARKET_SNAPSHOT_DATE = "2026-07-30"
MARKET_SNAPSHOT = {
    "CRM":  {"price": 180.71, "shares_m": 819.00},
    "ADBE": {"price": 247.90, "shares_m": 397.50},
    "NOW":  {"price": 110.07, "shares_m": 1030.00},
    "WDAY": {"price": 158.11, "shares_m": 246.97},
    "HUBS": {"price": 233.10, "shares_m": 51.22},
    "TEAM": {"price": 98.12,  "shares_m": 253.77},
    "DDOG": {"price": 268.56, "shares_m": 355.96},
    "SNOW": {"price": 298.10, "shares_m": 346.60},
    "ZS":   {"price": 148.39, "shares_m": 161.71},
    "CRWD": {"price": 185.22, "shares_m": 1020.00},
}

MILLION = 1_000_000

# --------------------------------------------------------------------------
# Formatting
# --------------------------------------------------------------------------

TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
LABEL_FONT = Font(size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # pale yellow -- editable
INPUT_FONT = Font(size=10, color="1F4E78", bold=True)
RESULT_FILL = PatternFill("solid", fgColor="D9E1F2")  # pale blue -- key output
RESULT_FONT = Font(size=10, bold=True)
THIN_BOTTOM = Border(bottom=Side(style="thin", color="BFBFBF"))

USD_M_FMT = "#,##0"
PCT_FMT = "0.0%"
USD_PER_SHARE_FMT = "$#,##0.00"


# --------------------------------------------------------------------------
# Loading cleaned data
# --------------------------------------------------------------------------

def load_companies() -> list:
    with open(PROCESSED_DATA_DIR / "companies.csv", newline="") as f:
        return list(csv.DictReader(f))


def load_financials_by_cik() -> dict:
    with open(PROCESSED_DATA_DIR / "dcf_financials.csv", newline="") as f:
        rows = list(csv.DictReader(f))
    by_cik = defaultdict(list)
    for r in rows:
        by_cik[r["cik"]].append(r)
    for cik in by_cik:
        by_cik[cik].sort(key=lambda r: int(r["fiscal_year"]))
    return by_cik


def to_float(val):
    return float(val) if val not in (None, "") else None


# --------------------------------------------------------------------------
# Deriving assumption defaults from history
# --------------------------------------------------------------------------

def compute_assumptions(records: list, ticker: str) -> dict:
    """Derive default model inputs from the cleaned historical records.
    Returns None (and prints a warning) if the company is missing data
    that's non-negotiable for a DCF (revenue, FCF, or shares outstanding
    in its most recent year)."""
    with_revenue = [r for r in records if to_float(r["revenue"]) is not None]
    if len(with_revenue) < 2:
        print(f"  [skip] {ticker}: fewer than 2 years of revenue history, can't compute a starting growth rate.")
        return None

    latest = with_revenue[-1]
    prior = with_revenue[-2]
    latest_rev = to_float(latest["revenue"])
    prior_rev = to_float(prior["revenue"])
    starting_growth = (latest_rev / prior_rev) - 1

    latest_fcf = to_float(latest["free_cash_flow"])
    if latest_fcf is None or latest_rev is None:
        print(f"  [skip] {ticker}: missing free_cash_flow or revenue in its most recent fiscal year.")
        return None

    margin_years = [r for r in with_revenue[-FCF_MARGIN_LOOKBACK_YEARS:]
                     if to_float(r["free_cash_flow"]) is not None]
    if not margin_years:
        print(f"  [skip] {ticker}: no recent year has both revenue and free_cash_flow.")
        return None
    fcf_margin = sum(to_float(r["free_cash_flow"]) / to_float(r["revenue"]) for r in margin_years) / len(margin_years)

    shares = to_float(latest["diluted_shares_outstanding"])
    if shares is None or shares == 0:
        print(f"  [skip] {ticker}: no diluted_shares_outstanding in its most recent fiscal year.")
        return None

    cash = to_float(latest["cash_and_equivalents"]) or 0.0
    debt = to_float(latest["total_debt"])
    if debt is None:
        debt = 0.0  # see module docstring: no debt tag reported this year -> treated as $0, not unknown.

    return {
        "base_fiscal_year": int(latest["fiscal_year"]),
        "base_revenue": latest_rev,
        "base_fcf": latest_fcf,
        "starting_growth": starting_growth,
        "fcf_margin": fcf_margin,
        "cash": cash,
        "debt": debt,
        "shares": shares,
        "history": with_revenue,  # for the historical-trend table
    }


# --------------------------------------------------------------------------
# Sheet building helpers
# --------------------------------------------------------------------------

def set_cell(ws, row, col, value, font=None, fill=None, number_format=None, border=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    if align:
        cell.alignment = align
    return cell


def build_sensitivity_formula(w_cell, g_cell, r0_cell, sg_cell, margin_cell,
                               debt_cell, cash_cell, shares_cell, years=PROJECTION_YEARS):
    """A single self-contained Excel formula computing implied value/share as
    a function of WACC (w_cell) and terminal growth (g_cell), with starting
    growth and FCF margin held fixed at the sheet's own assumption cells.

    This inlines the exact same fade/discount/terminal-value logic as the
    main projection above, just for one (WACC, terminal growth) point at a
    time -- necessary because growth_t depends on terminal growth too (the
    fade target), so revenue_t can't be pulled from the main projection rows
    when g varies independently across the grid. openpyxl also has no
    support for real Excel What-If Data Tables (the TABLE() array formula),
    so this closed-form expansion is what makes the grid live instead of a
    one-time snapshot of Python-computed numbers.
    """
    cum_growth_terms = []
    pv_terms = []
    final_fcf_expr = None
    for t in range(1, years + 1):
        growth_t = f"({sg_cell}+({g_cell}-{sg_cell})*{t}/{years})"
        cum_growth_terms.append(f"(1+{growth_t})")
        revenue_t_expr = f"{r0_cell}*" + "*".join(cum_growth_terms)
        fcf_t_expr = f"({revenue_t_expr})*{margin_cell}"
        pv_terms.append(f"(({fcf_t_expr})/(1+{w_cell})^{t})")
        if t == years:
            final_fcf_expr = fcf_t_expr

    tv_expr = f"(({final_fcf_expr})*(1+{g_cell}))/({w_cell}-{g_cell})"
    pv_tv_expr = f"(({tv_expr})/(1+{w_cell})^{years})"
    ev_expr = "+".join(pv_terms) + "+" + pv_tv_expr
    equity_expr = f"(({ev_expr})-{debt_cell}+{cash_cell})"
    return f"={equity_expr}/{shares_cell}"


def write_section_header(ws, row, text):
    set_cell(ws, row, 1, text, font=SECTION_FONT, border=THIN_BOTTOM)
    for col in range(2, 8):
        set_cell(ws, row, col, None, border=THIN_BOTTOM)
    return row + 1


# --------------------------------------------------------------------------
# Company sheet
# --------------------------------------------------------------------------

def build_company_sheet(wb, company: dict, assumptions: dict):
    ticker = company["ticker"]
    ws = wb.create_sheet(title=ticker)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14

    row = 1
    set_cell(ws, row, 1, f"{ticker} — {company['company_name']}", font=TITLE_FONT)
    row += 1
    set_cell(ws, row, 1, "Unlevered DCF | $ in millions except per-share and share count", font=LABEL_FONT)
    row += 2

    # ---- Assumptions -----------------------------------------------------
    row = write_section_header(ws, row, "Assumptions (edit the highlighted cells)")
    wacc_row = row
    set_cell(ws, row, 1, "WACC (discount rate)", font=LABEL_FONT)
    set_cell(ws, row, 2, DEFAULT_WACC, font=INPUT_FONT, fill=INPUT_FILL, number_format=PCT_FMT)
    row += 1
    tg_row = row
    set_cell(ws, row, 1, "Terminal growth rate", font=LABEL_FONT)
    set_cell(ws, row, 2, DEFAULT_TERMINAL_GROWTH, font=INPUT_FONT, fill=INPUT_FILL, number_format=PCT_FMT)
    row += 1
    sg_row = row
    set_cell(ws, row, 1, "Starting revenue growth rate", font=LABEL_FONT)
    set_cell(ws, row, 2, assumptions["starting_growth"], font=INPUT_FONT, fill=INPUT_FILL, number_format=PCT_FMT)
    set_cell(ws, row, 3, "default: most recent YoY revenue growth", font=Font(size=9, italic=True, color="808080"))
    row += 1
    margin_row = row
    set_cell(ws, row, 1, "Unlevered FCF margin (% of revenue)", font=LABEL_FONT)
    set_cell(ws, row, 2, assumptions["fcf_margin"], font=INPUT_FONT, fill=INPUT_FILL, number_format=PCT_FMT)
    set_cell(ws, row, 3, f"default: avg of last {FCF_MARGIN_LOOKBACK_YEARS} fiscal years' FCF / revenue",
             font=Font(size=9, italic=True, color="808080"))
    row += 2

    # ---- Historical trend --------------------------------------------------
    row = write_section_header(ws, row, "Historical Trend ($ in millions)")
    hist = assumptions["history"][-5:]  # up to 5 most recent actual years
    hist_year_row = row
    set_cell(ws, row, 1, "Fiscal year", font=LABEL_FONT)
    for i, r in enumerate(hist):
        set_cell(ws, row, 2 + i, int(r["fiscal_year"]), font=LABEL_FONT, align=Alignment(horizontal="right"))
    row += 1
    hist_rev_row = row
    set_cell(ws, row, 1, "Revenue", font=LABEL_FONT)
    for i, r in enumerate(hist):
        set_cell(ws, row, 2 + i, to_float(r["revenue"]) / MILLION, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1
    set_cell(ws, row, 1, "Revenue growth %", font=LABEL_FONT)
    for i in range(1, len(hist)):
        col = 2 + i
        prev_col_letter = get_column_letter(col - 1)
        col_letter = get_column_letter(col)
        formula = f"={col_letter}{hist_rev_row}/{prev_col_letter}{hist_rev_row}-1"
        set_cell(ws, row, col, formula, font=LABEL_FONT, number_format=PCT_FMT)
    row += 1
    hist_fcf_row = row
    set_cell(ws, row, 1, "Free cash flow", font=LABEL_FONT)
    for i, r in enumerate(hist):
        fcf = to_float(r["free_cash_flow"])
        set_cell(ws, row, 2 + i, fcf / MILLION if fcf is not None else None, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1
    set_cell(ws, row, 1, "FCF margin %", font=LABEL_FONT)
    for i, r in enumerate(hist):
        if to_float(r["free_cash_flow"]) is None:
            continue
        col_letter = get_column_letter(2 + i)
        formula = f"={col_letter}{hist_fcf_row}/{col_letter}{hist_rev_row}"
        set_cell(ws, row, 2 + i, formula, font=LABEL_FONT, number_format=PCT_FMT)
    row += 2

    # ---- Most recent balance sheet actuals --------------------------------
    row = write_section_header(ws, row, f"Balance Sheet Actuals — FY{assumptions['base_fiscal_year']} ($ in millions)")
    cash_row = row
    set_cell(ws, row, 1, "Cash & equivalents", font=LABEL_FONT)
    set_cell(ws, row, 2, assumptions["cash"] / MILLION, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1
    debt_row = row
    set_cell(ws, row, 1, "Total debt", font=LABEL_FONT)
    set_cell(ws, row, 2, assumptions["debt"] / MILLION, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1
    shares_row = row
    set_cell(ws, row, 1, "Diluted shares outstanding (millions)", font=LABEL_FONT)
    set_cell(ws, row, 2, assumptions["shares"] / MILLION, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 2

    # ---- Projection ---------------------------------------------------------
    row = write_section_header(ws, row, f"{PROJECTION_YEARS}-Year Projection ($ in millions)")
    base_year = assumptions["base_fiscal_year"]
    proj_first_col = 2  # column B

    year_row = row
    set_cell(ws, row, 1, "Fiscal year", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        set_cell(ws, row, proj_first_col + t - 1, f"FY{base_year + t}E", font=LABEL_FONT,
                 align=Alignment(horizontal="right"))
    row += 1

    growth_row = row
    set_cell(ws, row, 1, "Revenue growth rate", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        col = proj_first_col + t - 1
        # Linear fade: growth_t = starting_growth + (terminal_growth - starting_growth) * t/N
        # so growth_N == terminal growth exactly, matching the Gordon-growth terminal value.
        formula = (f"=$B${sg_row}+($B${tg_row}-$B${sg_row})*{t}/{PROJECTION_YEARS}")
        set_cell(ws, row, col, formula, font=LABEL_FONT, number_format=PCT_FMT)
    row += 1

    rev_row = row
    set_cell(ws, row, 1, "Projected revenue", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        col = proj_first_col + t - 1
        col_letter = get_column_letter(col)
        growth_cell = f"{col_letter}{growth_row}"
        if t == 1:
            # Chain off the historical trend table's last column (the base year's actual revenue).
            prior_rev_cell = f"{get_column_letter(1 + len(hist))}{hist_rev_row}"
        else:
            prior_rev_cell = f"{get_column_letter(col - 1)}{rev_row}"
        formula = f"={prior_rev_cell}*(1+{growth_cell})"
        set_cell(ws, row, col, formula, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1

    margin_ref_row = row
    set_cell(ws, row, 1, "FCF margin (assumption, held flat)", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        col = proj_first_col + t - 1
        set_cell(ws, row, col, f"=$B${margin_row}", font=LABEL_FONT, number_format=PCT_FMT)
    row += 1

    fcf_row = row
    set_cell(ws, row, 1, "Projected unlevered FCF", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        col = proj_first_col + t - 1
        col_letter = get_column_letter(col)
        formula = f"={col_letter}{rev_row}*{col_letter}{margin_ref_row}"
        set_cell(ws, row, col, formula, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1

    disc_row = row
    set_cell(ws, row, 1, "Discount factor", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        col = proj_first_col + t - 1
        formula = f"=1/(1+$B${wacc_row})^{t}"
        set_cell(ws, row, col, formula, font=LABEL_FONT, number_format="0.000")
    row += 1

    pv_fcf_row = row
    set_cell(ws, row, 1, "PV of FCF", font=LABEL_FONT)
    for t in range(1, PROJECTION_YEARS + 1):
        col = proj_first_col + t - 1
        col_letter = get_column_letter(col)
        formula = f"={col_letter}{fcf_row}*{col_letter}{disc_row}"
        set_cell(ws, row, col, formula, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 2

    # ---- Terminal value & valuation ----------------------------------------
    row = write_section_header(ws, row, "Terminal Value & Valuation")
    last_col_letter = get_column_letter(proj_first_col + PROJECTION_YEARS - 1)

    tv_row = row
    set_cell(ws, row, 1, f"Terminal value (end of FY{base_year + PROJECTION_YEARS})", font=LABEL_FONT)
    formula = f"={last_col_letter}{fcf_row}*(1+$B${tg_row})/($B${wacc_row}-$B${tg_row})"
    set_cell(ws, row, 2, formula, font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1

    pv_tv_row = row
    set_cell(ws, row, 1, "PV of terminal value", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=B{tv_row}*{last_col_letter}{disc_row}", font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1

    sum_pv_row = row
    set_cell(ws, row, 1, "Sum of PV of projected FCFs", font=LABEL_FONT)
    pv_fcf_range = f"B{pv_fcf_row}:{last_col_letter}{pv_fcf_row}"
    set_cell(ws, row, 2, f"=SUM({pv_fcf_range})", font=LABEL_FONT, number_format=USD_M_FMT)
    row += 1

    ev_row = row
    set_cell(ws, row, 1, "Enterprise value", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=B{pv_tv_row}+B{sum_pv_row}", font=RESULT_FONT, fill=RESULT_FILL, number_format=USD_M_FMT)
    row += 1

    set_cell(ws, row, 1, "Less: total debt", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=-B{debt_row}", font=LABEL_FONT, number_format=USD_M_FMT)
    less_debt_row = row
    row += 1

    set_cell(ws, row, 1, "Plus: cash & equivalents", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=B{cash_row}", font=LABEL_FONT, number_format=USD_M_FMT)
    plus_cash_row = row
    row += 1

    equity_row = row
    set_cell(ws, row, 1, "Equity value", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=B{ev_row}+B{less_debt_row}+B{plus_cash_row}",
             font=RESULT_FONT, fill=RESULT_FILL, number_format=USD_M_FMT)
    row += 1

    set_cell(ws, row, 1, "Diluted shares outstanding (millions)", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=B{shares_row}", font=LABEL_FONT, number_format=USD_M_FMT)
    shares_ref_row = row
    row += 1

    per_share_row = row
    set_cell(ws, row, 1, "Implied value per share", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=B{equity_row}/B{shares_ref_row}",
             font=RESULT_FONT, fill=RESULT_FILL, number_format=USD_PER_SHARE_FMT)
    row += 2

    # ---- Sensitivity grid ---------------------------------------------------
    # Fixed reference points for the closed-form formula below -- absolute so
    # every one of the 25 grid cells can reference the same source cells.
    r0_cell = f"${get_column_letter(1 + len(hist))}${hist_rev_row}"
    sg_cell = f"$B${sg_row}"
    margin_cell = f"$B${margin_row}"
    debt_cell = f"$B${debt_row}"
    cash_cell = f"$B${cash_row}"
    shares_cell = f"$B${shares_row}"

    row = write_section_header(ws, row, "Sensitivity: Implied Value/Share (WACC x Terminal Growth)")
    grid_header_row = row
    set_cell(ws, row, 2, "WACC \\ Term. g", font=Font(size=9, italic=True))
    for j, g_val in enumerate(SENSITIVITY_TG_VALUES):
        set_cell(ws, row, 3 + j, g_val, font=SECTION_FONT, number_format=PCT_FMT,
                 align=Alignment(horizontal="center"))
    row += 1

    grid_first_row = row
    for i, w_val in enumerate(SENSITIVITY_WACC_VALUES):
        r = row + i
        set_cell(ws, r, 2, w_val, font=SECTION_FONT, number_format=PCT_FMT)
        for j, g_val in enumerate(SENSITIVITY_TG_VALUES):
            col = 3 + j
            w_cell = f"$B${r}"
            g_cell = f"{get_column_letter(col)}${grid_header_row}"
            formula = build_sensitivity_formula(w_cell, g_cell, r0_cell, sg_cell, margin_cell,
                                                 debt_cell, cash_cell, shares_cell)
            set_cell(ws, r, col, formula, number_format=USD_PER_SHARE_FMT)
    grid_last_row = row + len(SENSITIVITY_WACC_VALUES) - 1
    last_tg_col_letter = get_column_letter(2 + len(SENSITIVITY_TG_VALUES))

    grid_range = f"C{grid_first_row}:{last_tg_col_letter}{grid_last_row}"
    ws.conditional_formatting.add(
        grid_range,
        ColorScaleRule(start_type="min", start_color="F8696B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B"),
    )
    row = grid_last_row + 2

    # ---- Tornado chart inputs -------------------------------------------------
    # Center column of the grid is the base-case terminal growth (3%); center
    # row is the base-case WACC (10%) -- so varying one axis while reading off
    # the grid's center row/column holds the other axis at the base case.
    center_tg_col_letter = get_column_letter(3 + len(SENSITIVITY_TG_VALUES) // 2)
    center_wacc_row = grid_first_row + len(SENSITIVITY_WACC_VALUES) // 2

    row = write_section_header(ws, row, "Tornado Inputs (other variable held at its base case)")
    tornado_header_row = row
    for i, h in enumerate(["Factor", "Low", "High", "Base (chart helper)", "Span (chart helper)"]):
        set_cell(ws, row, 1 + i, h, font=Font(size=9, italic=True))
    row += 1

    wacc_tornado_row = row
    set_cell(ws, row, 1, "WACC (8% → 12%)", font=LABEL_FONT)
    set_cell(ws, row, 2, f"={center_tg_col_letter}{grid_first_row}", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    set_cell(ws, row, 3, f"={center_tg_col_letter}{grid_last_row}", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    set_cell(ws, row, 4, f"=MIN(B{row},C{row})", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    set_cell(ws, row, 5, f"=ABS(C{row}-B{row})", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    row += 1

    tg_tornado_row = row
    set_cell(ws, row, 1, "Terminal growth (2% → 4%)", font=LABEL_FONT)
    set_cell(ws, row, 2, f"=C{center_wacc_row}", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    set_cell(ws, row, 3, f"={last_tg_col_letter}{center_wacc_row}", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    set_cell(ws, row, 4, f"=MIN(B{row},C{row})", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    set_cell(ws, row, 5, f"=ABS(C{row}-B{row})", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
    row += 2

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Value/Share Sensitivity (Tornado)"
    chart.y_axis.title = None
    chart.x_axis.title = "Implied value/share ($)"
    chart.height = 6
    chart.width = 16
    chart.legend = None

    cats = Reference(ws, min_col=1, min_row=wacc_tornado_row, max_row=tg_tornado_row)
    base_ref = Reference(ws, min_col=4, min_row=tornado_header_row, max_row=tg_tornado_row)
    span_ref = Reference(ws, min_col=5, min_row=tornado_header_row, max_row=tg_tornado_row)
    chart.add_data(base_ref, titles_from_data=True)
    chart.add_data(span_ref, titles_from_data=True)
    chart.set_categories(cats)

    # First (base) series is a spacer, not real data -- keep it invisible so
    # the chart reads as floating bars rather than stacked-from-zero bars.
    chart.series[0].graphicalProperties = GraphicalProperties()
    chart.series[0].graphicalProperties.noFill = True
    chart.series[1].graphicalProperties = GraphicalProperties(solidFill="4472C4")

    ws.add_chart(chart, f"A{row}")

    return {
        "wacc_cell": f"'{ticker}'!B{wacc_row}",
        "terminal_growth_cell": f"'{ticker}'!B{tg_row}",
        "revenue_cell": f"'{ticker}'!{get_column_letter(1 + len(hist))}{hist_rev_row}",
        "ev_cell": f"'{ticker}'!B{ev_row}",
        "equity_value_cell": f"'{ticker}'!B{equity_row}",
        "shares_cell": f"'{ticker}'!B{shares_ref_row}",
        "per_share_cell": f"'{ticker}'!B{per_share_row}",
        "low_estimate_cell": f"'{ticker}'!C{grid_last_row}",
        "high_estimate_cell": f"'{ticker}'!{last_tg_col_letter}{grid_first_row}",
    }


# --------------------------------------------------------------------------
# Summary sheet
# --------------------------------------------------------------------------

def build_summary_sheet(wb, rows: list):
    """rows: list of (company_dict, cell_refs) in the order sheets were built."""
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    for col in "CDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 16
    for col in ("F", "J", "K", "M"):  # headers here run longer than the default width
        ws.column_dimensions[col].width = 22

    set_cell(ws, 1, 1, "Tech DCF Valuation — Summary", font=TITLE_FONT)
    set_cell(ws, 2, 1, "Every column here is a live link into that company's own sheet — "
                        "edit assumptions on a company tab and this table updates.",
             font=Font(size=9, italic=True, color="808080"))

    headers = ["Ticker", "Company", "WACC", "Terminal growth",
               f"FY revenue ($M)", "Enterprise value ($M)", "Equity value ($M)",
               "Diluted shares (M)", "Implied value/share",
               "Low est. (12% WACC, 2% g)", "High est. (8% WACC, 4% g)",
               "Price (input)", "Shares O/S, M (input)",
               "Market cap ($M)", "DCF vs. market"]
    header_row = 4
    for i, h in enumerate(headers):
        set_cell(ws, header_row, 1 + i, h, font=SECTION_FONT, border=THIN_BOTTOM)
    set_cell(ws, 3, 1, f"\"Price (input)\" / \"Shares O/S\" default to a market close snapshot from "
                        f"{MARKET_SNAPSHOT_DATE} -- edit those two highlighted cells any time to "
                        f"re-run the comparison with a current quote.",
             font=Font(size=9, italic=True, color="808080"))

    row = header_row + 1
    for company, refs in rows:
        ticker = company["ticker"]
        snapshot = MARKET_SNAPSHOT.get(ticker, {})
        set_cell(ws, row, 1, ticker, font=LABEL_FONT)
        set_cell(ws, row, 2, company["company_name"], font=LABEL_FONT)
        set_cell(ws, row, 3, f"={refs['wacc_cell']}", font=LABEL_FONT, number_format=PCT_FMT)
        set_cell(ws, row, 4, f"={refs['terminal_growth_cell']}", font=LABEL_FONT, number_format=PCT_FMT)
        set_cell(ws, row, 5, f"={refs['revenue_cell']}", font=LABEL_FONT, number_format=USD_M_FMT)
        set_cell(ws, row, 6, f"={refs['ev_cell']}", font=LABEL_FONT, number_format=USD_M_FMT)
        set_cell(ws, row, 7, f"={refs['equity_value_cell']}", font=LABEL_FONT, number_format=USD_M_FMT)
        set_cell(ws, row, 8, f"={refs['shares_cell']}", font=LABEL_FONT, number_format=USD_M_FMT)
        set_cell(ws, row, 9, f"={refs['per_share_cell']}", font=RESULT_FONT, fill=RESULT_FILL,
                 number_format=USD_PER_SHARE_FMT)
        set_cell(ws, row, 10, f"={refs['low_estimate_cell']}", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
        set_cell(ws, row, 11, f"={refs['high_estimate_cell']}", font=LABEL_FONT, number_format=USD_PER_SHARE_FMT)
        set_cell(ws, row, 12, snapshot.get("price"), font=INPUT_FONT, fill=INPUT_FILL, number_format=USD_PER_SHARE_FMT)
        set_cell(ws, row, 13, snapshot.get("shares_m"), font=INPUT_FONT, fill=INPUT_FILL, number_format=USD_M_FMT)
        set_cell(ws, row, 14, f"=L{row}*M{row}", font=LABEL_FONT, number_format=USD_M_FMT)
        # Market cap and DCF equity value are both share-count-invariant totals,
        # so this comparison holds even for a company (CRWD) whose share count
        # changed between its FY-end 10-K and this snapshot due to a split.
        set_cell(ws, row, 15, f"=(G{row}-N{row})/N{row}", font=RESULT_FONT, fill=RESULT_FILL, number_format=PCT_FMT)
        row += 1


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    companies = load_companies()
    financials = load_financials_by_cik()

    wb = Workbook()
    wb.remove(wb.active)  # replaced by the Summary sheet at index 0

    summary_rows = []
    for company in companies:
        ticker = company["ticker"]
        records = financials.get(company["cik"], [])
        print(f"[model] {ticker} ...", end=" ")
        assumptions = compute_assumptions(records, ticker)
        if assumptions is None:
            continue
        refs = build_company_sheet(wb, company, assumptions)
        summary_rows.append((company, refs))
        print(f"FY{assumptions['base_fiscal_year']} base, "
              f"{assumptions['starting_growth']*100:.1f}% starting growth, "
              f"{assumptions['fcf_margin']*100:.1f}% FCF margin")

    build_summary_sheet(wb, summary_rows)
    wb.save(OUTPUT_PATH)
    print(f"\nDone. Wrote {len(summary_rows)} company sheets + Summary to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
